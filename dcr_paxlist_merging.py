import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
import traceback
from datetime import datetime
import os

# Konfigurasi halaman
st.set_page_config(
    page_title="DCR-PAXLIST Automation",
    page_icon="🦁",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS untuk styling
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #1e3c72 0%, #2a5298 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
    .error-box {
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        color: #721c24;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #cce7ff;
        border: 1px solid #99d5ff;
        color: #004080;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

def validate_paxlist(df):
    """Validasi struktur file PAXLIST"""
    required_columns = ["Crew ID", "Booking Code"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        return False, f"Kolom yang hilang: {', '.join(missing_columns)}"
    
    # Cek apakah ada data
    if df.empty:
        return False, "File PAXLIST kosong"
    
    # Cek apakah Crew ID berisi data yang valid
    if df["Crew ID"].isna().all():
        return False, "Kolom Crew ID tidak memiliki data yang valid"
    
    return True, "Valid"

def validate_dcr(df):
    """Validasi struktur file DCR"""
    required_columns = ["CREW LIST"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        return False, f"Kolom yang hilang: {', '.join(missing_columns)}"
    
    if df.empty:
        return False, "File DCR kosong"
    
    return True, "Valid"

def extract_crew_id(crew_list_value):
    """Ekstrak Crew ID dari kolom CREW LIST"""
    if pd.isna(crew_list_value):
        return None
    
    crew_str = str(crew_list_value).strip()
    if "/" in crew_str:
        crew_id = crew_str.split("/")[0].strip()
    else:
        crew_id = crew_str
    
    try:
        return int(crew_id)
    except (ValueError, TypeError):
        return None

def group_booking_codes(paxlist_df):
    """Kelompokkan booking codes berdasarkan Crew ID"""
    # Bersihkan data
    paxlist_clean = paxlist_df.dropna(subset=["Crew ID", "Booking Code"])
    
    # Konversi Crew ID ke numeric
    paxlist_clean["Crew ID"] = pd.to_numeric(paxlist_clean["Crew ID"], errors="coerce")
    paxlist_clean = paxlist_clean.dropna(subset=["Crew ID"])
    
    # Kelompokkan booking codes
    grouped = paxlist_clean.groupby("Crew ID")["Booking Code"].apply(
        lambda x: ", ".join(x.astype(str).unique())
    ).reset_index()
    
    return grouped

def apply_formatting(file_path, booking_column_name="Booking Code"):
    """Terapkan formatting pada file Excel"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Cari kolom booking code
    booking_col = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == booking_column_name:
            booking_col = i
            break
    
    if booking_col:
        # Style untuk header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Apply header styling
        ws.cell(row=1, column=booking_col).fill = header_fill
        ws.cell(row=1, column=booking_col).font = header_font
        
        # Style untuk JUMPSEAT
        jumpseat_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        jumpseat_font = Font(color="FFFFFF", bold=True)
        
        # Terapkan formatting untuk baris yang mengandung JUMPSEAT
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=booking_col)
            if cell.value and "JUMPSEAT" in str(cell.value).upper():
                cell.fill = jumpseat_fill
                cell.font = jumpseat_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
    return True

def process_files(paxlist_file, dcr_file, dcr_header_row=1):
    """Proses utama penggabungan file"""
    try:
        # Baca file PAXLIST
        if paxlist_file.name.endswith('.csv'):
            paxlist_df = pd.read_csv(paxlist_file)
        else:
            paxlist_df = pd.read_excel(paxlist_file)
        
        # Baca file DCR
        dcr_df = pd.read_excel(dcr_file, header=dcr_header_row)
        
        # Validasi file
        pax_valid, pax_msg = validate_paxlist(paxlist_df)
        if not pax_valid:
            return None, f"Error PAXLIST: {pax_msg}"
        
        dcr_valid, dcr_msg = validate_dcr(dcr_df)
        if not dcr_valid:
            return None, f"Error DCR: {dcr_msg}"
        
        # Kelompokkan booking codes
        paxlist_grouped = group_booking_codes(paxlist_df)
        
        # Ekstrak Crew ID dari DCR
        dcr_df["Crew ID"] = dcr_df["CREW LIST"].apply(extract_crew_id)
        
        # Merge data
        dcr_merged = dcr_df.merge(paxlist_grouped, on="Crew ID", how="left")
        
        # Isi nilai kosong dengan "-"
        dcr_merged["Booking Code"] = dcr_merged["Booking Code"].fillna("-")
        
        return dcr_merged, "Success"
        
    except Exception as e:
        return None, f"Error: {str(e)}\n\nDetail:\n{traceback.format_exc()}"

# Header aplikasi
st.markdown("""
<div class="main-header">
    <h1 style="color: white;">🦁 DCR-PAXLIST Automation Tool</h1>
    <p>Otomatisasi penggabungan data Booking Code dari PAXLIST ke DCR</p>
</div>
""", unsafe_allow_html=True)

# Sidebar untuk instruksi
with st.sidebar:
    st.header("📋 Instruksi Penggunaan")
    st.markdown("""
    **Langkah-langkah:**
    1. Upload file DCR (.xlsx)
    2. Upload file PAXLIST (.xlsx/.csv)
    3. Sesuaikan pengaturan jika diperlukan
    4. Klik 'Proses File'
    5. Download hasil yang sudah diproses
    
    **Persyaratan File:**
    - **DCR**: Harus memiliki kolom 'CREW LIST'
    - **PAXLIST**: Harus memiliki kolom 'Crew ID' dan 'Booking Code'
    """)
    
    st.header("⚙️ Pengaturan")
    dcr_header_row = st.selectbox(
        "Baris header DCR:",
        options=[0, 1, 2],
        index=1,
        help="Pilih baris mana yang menjadi header di file DCR (0 = baris pertama)"
    )
    
    apply_formatting_option = st.checkbox(
        "Terapkan formatting",
        value=True,
        help="Warnai cell yang mengandung 'JUMPSEAT' dengan merah"
    )

# Main content
col1, col2 = st.columns(2)

with col1:
    st.subheader("📁 Upload File DCR")
    dcr_file = st.file_uploader(
        "Pilih file DCR (.xlsx)",
        type=['xlsx'],
        key="dcr_file",
        help="File DCR yang akan menjadi template utama"
    )
    
    if dcr_file:
        st.success(f"✅ DCR file uploaded: {dcr_file.name}")

with col2:
    st.subheader("📊 Upload File PAXLIST")
    paxlist_file = st.file_uploader(
        "Pilih file PAXLIST (.xlsx/.csv)",
        type=['xlsx', 'csv'],
        key="paxlist_file",
        help="File yang berisi data booking code crew"
    )
    
    if paxlist_file:
        st.success(f"✅ PAXLIST file uploaded: {paxlist_file.name}")

# Preview data
if paxlist_file:
    with st.expander("👀 Preview Data PAXLIST"):
        try:
            if paxlist_file.name.endswith('.csv'):
                preview_df = pd.read_csv(paxlist_file)
            else:
                preview_df = pd.read_excel(paxlist_file)
            
            st.dataframe(preview_df.head(10))
            st.info(f"Total baris: {len(preview_df)}")
        except Exception as e:
            st.error(f"Error reading preview: {str(e)}")

if dcr_file:
    with st.expander("👀 Preview Data DCR"):
        try:
            preview_df = pd.read_excel(dcr_file, header=dcr_header_row)
            st.dataframe(preview_df.head(10))
            st.info(f"Total baris: {len(preview_df)}")
        except Exception as e:
            st.error(f"Error reading preview: {str(e)}")

# Tombol proses
if dcr_file and paxlist_file:
    if st.button("🚀 Proses File", type="primary", use_container_width=True):
        with st.spinner("Memproses file..."):
            result_df, message = process_files(paxlist_file, dcr_file, dcr_header_row)
            
            if result_df is not None:
                st.markdown('<div class="success-box">✅ File berhasil diproses!</div>', unsafe_allow_html=True)
                
                # Statistik hasil
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Crew", len(result_df))
                with col2:
                    booking_filled = len(result_df[result_df["Booking Code"] != "-"])
                    st.metric("Crew dengan Booking", booking_filled)
                with col3:
                    jumpseat_count = len(result_df[result_df["Booking Code"].str.contains("JUMPSEAT", case=False, na=False)])
                    st.metric("JUMPSEAT Bookings", jumpseat_count)
                
                # Preview hasil
                with st.expander("👀 Preview Hasil"):
                    st.dataframe(result_df)
                
                # Simpan ke buffer untuk download
                output_buffer = io.BytesIO()
                
                # Simpan ke Excel
                result_df.to_excel(output_buffer, index=False)
                
                if apply_formatting_option:
                    # Terapkan formatting
                    output_buffer.seek(0)
                    temp_path = f"temp_dcr_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    
                    with open(temp_path, 'wb') as f:
                        f.write(output_buffer.getvalue())
                    
                    try:
                        apply_formatting(temp_path)
                        
                        # Baca kembali file yang sudah diformat
                        with open(temp_path, 'rb') as f:
                            formatted_data = f.read()
                        
                        # Hapus file temp
                        os.remove(temp_path)
                        
                        # Download button
                        filename = f"DCR_Updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        st.download_button(
                            label="📥 Download DCR Hasil (dengan formatting)",
                            data=formatted_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    except Exception as format_error:
                        st.warning(f"Formatting error: {str(format_error)}")
                        # Fallback ke download tanpa formatting
                        output_buffer.seek(0)
                        filename = f"DCR_Updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        st.download_button(
                            label="📥 Download DCR Hasil (tanpa formatting)",
                            data=output_buffer.getvalue(),
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                else:
                    # Download tanpa formatting
                    output_buffer.seek(0)
                    filename = f"DCR_Updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    st.download_button(
                        label="📥 Download DCR Hasil",
                        data=output_buffer.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            else:
                st.markdown(f'<div class="error-box">❌ {message}</div>', unsafe_allow_html=True)

# Footer
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #666; padding: 1rem;">
    <p>DCR-PAXLIST Automation Tool | Dibuat dengan menggunakan Streamlit</p>
    <p><small>© Maurino Audrian Putra</small></p>
</div>
""", unsafe_allow_html=True)